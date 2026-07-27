"""
모바일라우터 사용량 추출 -> 정리 -> 기입 자동화
IPR-400 NMS (http://nms.iproad.co.kr/ipr-400/dashboard)

===============================================================
설치 (실행 전 딱 한 번만 하면 됨)
===============================================================
1. 파이썬 패키지 설치 (명령 프롬프트/터미널에서 실행)
     pip install selenium openpyxl xlwings

   * "pip"가 안 먹으면: python -m pip install selenium openpyxl xlwings

2. 컴퓨터에 아래 두 프로그램이 설치되어 있어야 함
     - Google Chrome        (NMS 로그인/다운로드용)
     - Microsoft Excel       (xb 체크리스트의 순환수식 계산용)

3. 크롬드라이버는 별도 설치 불필요
     (Selenium 4.6+ 는 Selenium Manager가 자동으로 관리함)

4. XB_PATH를 실제 xb 엑셀 파일 경로로 채워넣을 것
     (윈도우 경로 예: "C:\\Users\\사용자명\\Documents\\체크리스트.xlsx")
===============================================================

전체 흐름
    1) login()                 - NMS 로그인 (보안문자는 직접 입력, 나머지는 자동)
    2) download_router_status() - 전체 라우터 상태정보 엑셀(xa) 다운로드
    3) organize_sheet()         - xa의 D열(상세위치)을 정리해서 '새탭'에 MR 번호순 정렬
    4) fill_xb_from_xa()        - xa의 사용량 값을 xb(누적 체크리스트)에 기입 + 계산

로그인 방식: Selenium (실제 브라우저 제어)
    로그인 시 '보안문자'(캡차)가 매번 나와서 자동으로 풀거나 우회하지 않고,
    ID/PW만 자동 입력한 뒤 보안문자는 사용자가 직접 입력하도록 함.
    보안문자 입력이 끝나면(3초간 값 변화 없으면) 자동으로 로그인 버튼을 누름.

xb 기입 방식: xlwings (실제 엑셀 앱 제어)
    xb의 today/Fixed daily row 탭은 자기 자신을 참조하는 순환수식으로 되어있어서
    (M2 상태가 바뀌는 '그 순간'에 값을 스냅샷 떠야 함) openpyxl로 셀 값만 고쳐서는
    정확히 계산되지 않음. 그래서 xlwings로 실제 엑셀을 열어 today!M2를
    TYPING -> START -> INPUT -> START 순서로 바꾸며 매번 재계산을 강제로 시킴.
"""

import os
import time

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ===================== 사용자 설정 (여기만 필요하면 수정) =====================
USER_ID = "jcdkorea"
USER_PW = "jcdkorea00!"
LOGIN_URL = "http://nms.iproad.co.kr/ipr-400/dashboard"

# 엑셀 다운로드 파일이 저장될 폴더 (스크립트와 같은 위치의 downloads 폴더)
DOWNLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "downloads")

# 누적 기입할 체크리스트 엑셀(xb) 경로 (여기에 직접 채워넣기)
XB_PATH = ""
# ==============================================================================


# ---------------------------------------------------------------------------
# 1) 로그인
# ---------------------------------------------------------------------------
def login():
    """NMS에 로그인해서 대시보드까지 접속된 driver를 반환함"""
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    # 다운로드 시 항상 DOWNLOAD_DIR에 자동 저장되도록 설정 (다운로드 위치 물어보는 팝업 방지)
    # + 세이프 브라우징이 xlsx를 '안전하지 않은 다운로드'로 막는 것 방지
    options = webdriver.ChromeOptions()
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": DOWNLOAD_DIR,
            "download.prompt_for_download": False,
            "safebrowsing.enabled": False,
            "safebrowsing.disable_download_protection": True,
        },
    )
    options.add_argument("--safebrowsing-disable-download-protection")

    driver = webdriver.Chrome(options=options)

    # '안전하지 않은 다운로드가 차단됨' 알림은 페이지 요소가 아니라 크롬 자체 UI라
    # Selenium으로 클릭할 수 없음. 대신 DevTools Protocol로 다운로드를
    # 무조건 허용하도록 설정해서 그 알림 자체가 안 뜨게 만듦.
    driver.execute_cdp_cmd(
        "Page.setDownloadBehavior",
        {"behavior": "allow", "downloadPath": DOWNLOAD_DIR},
    )

    driver.maximize_window()
    driver.get(LOGIN_URL)

    wait = WebDriverWait(driver, 15)

    # ID / 비밀번호 입력란 (실제 페이지 HTML의 id 속성 기준)
    id_input = wait.until(EC.presence_of_element_located((By.ID, "userID")))
    pw_input = driver.find_element(By.ID, "password")

    id_input.clear()
    id_input.send_keys(USER_ID)
    pw_input.clear()
    pw_input.send_keys(USER_PW)

    # 보안문자를 틀렸을 때 다시 시도할 수 있도록 전체를 반복문으로 감쌈
    while True:
        # 보안문자 입력란(answer)에 바로 입력할 수 있도록 커서(포커스) 이동
        # (재시도 시 새 이미지로 answer 엘리먼트가 다시 그려질 수 있어 매번 새로 찾음)
        answer_input = driver.find_element(By.ID, "answer")
        answer_input.clear()
        answer_input.click()

        print("브라우저 창에서 보안문자만 입력해줘. 입력을 마치면 자동으로 로그인돼.")

        # 값이 채워지고, 일정 시간(3초) 동안 더 이상 바뀌지 않으면
        # (=입력 완료로 판단) 자동으로 로그인 버튼 클릭
        last_value = ""
        stable_since = None
        while True:
            time.sleep(0.2)
            current_value = answer_input.get_attribute("value")
            if current_value != last_value:
                last_value = current_value
                stable_since = time.time()
                continue
            if current_value.strip() and stable_since and (time.time() - stable_since) >= 3:
                break

        driver.find_element(By.ID, "frmSubmit").click()

        # 보안문자가 틀리면 브라우저 기본 alert 팝업("보안 문자를 확인하세요")이 뜸
        # 팝업이 뜨는지 3초간 확인 -> 뜨면 확인 누르고 다시 보안문자 입력 대기
        try:
            WebDriverWait(driver, 3).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            print(f"로그인 실패: {alert.text}")
            alert.accept()  # 팝업의 '확인' 버튼 클릭
            continue  # 로그인 화면으로 돌아가 보안문자 다시 입력
        except TimeoutException:
            pass  # 팝업 없음 -> 로그인 성공으로 진행

        # 로그인 성공 여부 확인
        wait.until(EC.url_contains("dashboard"))
        print("로그인 성공, 대시보드 접속 완료")
        break

    return driver


# ---------------------------------------------------------------------------
# 2) 엑셀(xa) 다운로드
# ---------------------------------------------------------------------------
def download_router_status(driver):
    """로그인 후 '라우터 상태정보' 메뉴로 이동해서 전체 라우터 상태정보 엑셀 다운로드
    다운로드가 끝난 파일의 경로를 반환함"""
    wait = WebDriverWait(driver, 15)

    # 라우터 상태정보 메뉴 클릭
    menu_link = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='modemall']"))
    )
    menu_link.click()

    # 전체 라우터 상태정보 엑셀 다운로드 버튼 클릭
    download_button = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(@onclick, 'modemallexceldownload')]")
        )
    )

    before_files = set(os.listdir(DOWNLOAD_DIR))
    download_button.click()
    print(f"엑셀 다운로드 시작. 저장 위치: {DOWNLOAD_DIR}")

    # 새 파일이 생기고, 다운로드 중 임시파일(.crdownload)이 아닐 때까지 대기
    deadline = time.time() + 30
    new_file = None
    while time.time() < deadline:
        new_files = {
            f for f in set(os.listdir(DOWNLOAD_DIR)) - before_files
            if not f.endswith(".crdownload")
        }
        if new_files:
            new_file = new_files.pop()
            break
        time.sleep(0.5)

    if not new_file:
        raise TimeoutError("다운로드된 파일을 찾지 못했어. downloads 폴더를 확인해줘.")

    filepath = os.path.join(DOWNLOAD_DIR, new_file)
    print(f"다운로드 완료: {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# 3) xa 정리 (D열 나누기 + MR 번호순 정렬)
# ---------------------------------------------------------------------------
def organize_sheet(filepath):
    """다운로드한 엑셀의 D열(상세위치, 예: '91. 예술의 전당')을
    '.' 기준으로 나눠서 '새탭' 시트에 정리함.
    - 원본 데이터는 그대로 두고 D열 뒤에 나눠진 조각을 넣을 컬럼을 추가함 (기존 데이터 안 밀림)
    - 맨 앞 숫자(위치번호) 기준 오름차순 정렬
    - 맨 앞 숫자에 'MR ' 접두어를 붙임 (2자리 기준 0채움, 예: 1 -> MR 01, 91 -> MR 91)
    """
    wb = openpyxl.load_workbook(filepath)
    src = wb.active  # 원본 데이터 시트 (다운로드된 그대로)

    rows = list(src.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]

    # D열(인덱스 3)을 '.'으로 나눴을 때 최대 몇 조각까지 나오는지 확인
    max_parts = max(len(str(row[3]).split(".")) for row in data_rows)
    extra_cols = max_parts - 1  # D열 뒤에 추가로 필요한 컬럼 수

    # 새 탭 생성 (이미 있으면 지우고 새로 만듦)
    if "새탭" in wb.sheetnames:
        del wb["새탭"]
    new_ws = wb.create_sheet("새탭")

    # 헤더 구성: A~C(기존) + D(번호) + E(상세위치, 나머지 조각용 빈 헤더) + 기존 E~ (한 칸씩 밀림)
    new_header = (
        list(header[:3])
        + ["번호(MR)", "상세위치"]
        + [""] * (extra_cols - 1)
        + list(header[4:])
    )
    new_ws.append(new_header)

    processed_rows = []
    for row in data_rows:
        parts = str(row[3]).split(".")
        parts += [""] * (max_parts - len(parts))  # 조각 수가 모자라면 빈 값으로 채움

        number = int(parts[0].strip())  # 정렬 및 MR 라벨링 기준이 되는 맨 앞 숫자

        new_row = list(row[:3]) + parts + list(row[4:])
        processed_rows.append((number, new_row))

    # 맨 앞 숫자 기준 오름차순 정렬
    processed_rows.sort(key=lambda item: item[0])

    for number, new_row in processed_rows:
        new_row[3] = f"MR {number:02d}"  # 맨 앞 숫자에 MR 접두어 붙이기
        new_ws.append(new_row)

    wb.save(filepath)
    print(f"'새탭' 시트에 정리 완료: {filepath}")


# ---------------------------------------------------------------------------
# 4) xb(누적 체크리스트)에 기입 + 계산
# ---------------------------------------------------------------------------
def build_xa_usage_map(xa_filepath):
    """xa(정렬된 엑셀)의 '새탭' 시트에서 D열(번호 MR) -> M열(당월 데이터 사용량) 매핑을 만듦"""
    wb = openpyxl.load_workbook(xa_filepath, data_only=True)
    ws = wb["새탭"]

    usage_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mr_code = row[3]  # D열: 번호(MR)
        usage = row[12]  # M열: 당월 데이터 사용량(MByte)
        if mr_code:
            usage_map[str(mr_code).strip()] = usage
    return usage_map


def fill_xb_from_xa(xa_filepath, xb_filepath=XB_PATH):
    """
    xb(누적 체크리스트) 엑셀의 today 탭에 xa의 당월 사용량 값을 채우고,
    M2 드롭박스를 TYPING -> START -> INPUT -> START 순서로 바꿔가며 실제 계산을 진행함.

    today 탭의 G/H/I열과 Fixed daily row 탭의 날짜별 열은 자기 자신을 참조하는
    순환수식으로 되어있어서(M2 상태가 바뀌는 '그 순간'에 스냅샷을 떠야 함),
    openpyxl로 파일만 고쳐서는 정확한 계산이 안 됨. 그래서 xlwings로 실제
    엑셀 앱을 열어 각 단계마다 재계산을 강제로 시킴. xb_filepath에 바로 덮어씀.
    """
    import xlwings as xw

    usage_map = build_xa_usage_map(xa_filepath)

    app = xw.App(visible=True)

    # 1. 반복 계산(순환 참조) 활성화 - 작업 전 필수 체크사항
    #    (엑셀 환경설정 > 수식 계산 > 반복 계산 사용)
    try:
        app.api.Iteration = True
        app.api.MaxIterations = 100
        app.api.MaxChange = 0.001
        print("반복 계산(순환 참조) 활성화 완료")
    except Exception as e:
        print(
            f"반복 계산 자동 설정 실패({e}). "
            "엑셀 환경설정 > 수식 계산 > 반복 계산 사용을 직접 체크한 뒤 다시 실행해줘."
        )

    wb = app.books.open(xb_filepath)
    ws = wb.sheets["today"]

    def set_switch(value):
        ws.range("M2").value = value
        app.calculate()  # 강제 재계산 -> 이 순간의 값을 순환수식이 스냅샷

    # 2. TYPING으로 변경 (H열이 이전 G열 값을 스냅샷)
    set_switch("TYPING")

    # 3. xa 새탭 M열 값을, C열(MR 코드)로 매칭해서 K열에 채움
    last_row = ws.range("C1").end("down").row
    mr_codes = ws.range(f"C2:C{last_row}").value

    k_values = []
    not_found = []
    for mr_code in mr_codes:
        key = str(mr_code).strip() if mr_code else None
        if key in usage_map:
            k_values.append(usage_map[key])
        else:
            k_values.append(None)
            not_found.append(mr_code)

    ws.range(f"K2:K{last_row}").options(transpose=True).value = k_values

    if not_found:
        print(f"xa에서 매칭되는 값을 못 찾은 MR 코드: {not_found}")

    # 4. START -> INPUT -> START 순서로 전환 (매번 재계산)
    set_switch("START")
    set_switch("INPUT")  # 이 시점에 Fixed daily row에 오늘 날짜 값이 기록됨
    set_switch("START")

    wb.save()
    print(f"작업 완료. 엑셀에서 결과를 직접 확인해줘: {xb_filepath}")
    # 결과를 바로 확인할 수 있도록 앱/파일은 닫지 않고 열어둠

    return xb_filepath


# ---------------------------------------------------------------------------
# 실행
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    driver = login()
    xa_filepath = download_router_status(driver)
    organize_sheet(xa_filepath)

    time.sleep(2)  # 다운로드/정리 마무리 대기 후 바로 다음 단계 진행
    driver.quit()

    fill_xb_from_xa(xa_filepath)
